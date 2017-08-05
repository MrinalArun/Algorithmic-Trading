#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 24 12:22:15 2016

@Group Project - An Efficient Tracking Of Volume Weighted Average Price And Its Implementation
Algorithmic Trading 
Libraries used
openpyxl - to read and write into excel files

"""
from __future__ import division
from datetime import datetime
import openpyxl as excel
import os
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

#Class manages Price Volume Time data for a single stock for a single day
class PriceVolumeTime:
    CONST_LEN = 390
    # Default Initializer for all the variables.
    def __init__(self,stockname = None, xdate = None):
        self.date = xdate
        self.stockname = stockname
        self.price = []
        self.volume = []
        self.time = []
    
    def VWAP(self):
        index =0
        PV =0;
        if not self.time:
            return 0
        if len(self.price) != PriceVolumeTime.CONST_LEN:            
            print 'Problem with the number of data sets in %s Stock for day %s' % (self.stockname , str(self.date))
        for index in range(len(self.volume)):
            PV = PV + self.price[index]*self.volume[index]
        return PV/sum(self.volume)
    def TotalVol(self):
        return sum(self.volume)
    def Length(self):
        if len(self.price) == len(self.time) and len(self.time) == len(self.volume):
            return len(self.price)
        else:
            print('Exception: Length of Price, Vol, Time Vectors does not match in %s %s' % (self.date , self.stockname))

# Class manages the Price Volume Time data for a single stock for multiple days
class PVTforAllDays:
    def __init__(self,stockname = None):
        self.PriceVolumeTimeVectors = []
        self.stockname = stockname
    def NumberofDays(self):
        return len(self.PriceVolumeTimeVectors)
    def AppendDay(self,PriceVolumeTimeVector):
        self.PriceVolumeTimeVectors.append(PriceVolumeTimeVector)
    def __getitem__(self,key):
        return self.PriceVolumeTimeVectors[key]
    def Stockname(self):
        return self.stockname

# Returns the final Ut vector Args(Main Data struct, starting stock count,ending stock count, starting day count, ending day count)
def ReturnUt(Object_stocks,stock_start,stock_end,day_start,day_end):
    if stock_end > len(Object_stocks):
        print 'Not enough stocks in Data'
        return 0
    min_days = 10000
    stock_1 = ""
    for x in Object_stocks:
        if min_days > x.NumberofDays():
            min_days = x.NumberofDays()
            stock_1 = x.Stockname()
    if day_end > min_days:
        print 'Not enough days data in one stock:'
        print (stock_1)
        return 0
    
    z = stock_start
    
    Ut = []
    SumMt = []
    for x in range(390):
        Ut.append(0)
        SumMt.append(0)
    Mt = []
    Tempt = []
    index = day_start    
    while (index <= day_end):
        while(z <= stock_end):
            Mt = Object_stocks[z].PriceVolumeTimeVectors[index].volume          
            for x in range(len(Mt)):
                Tempt.append(Mt[x]/Object_stocks[z].PriceVolumeTimeVectors[index].TotalVol())        
            for x in range(len(Tempt)):
                SumMt[x] = SumMt[x] + Tempt[x]            
            del Tempt[:]
            z += 1
        z = stock_start
        index += 1
    for x in range(len(SumMt)):
        Ut[x] = SumMt[x]/((stock_end-stock_start+1)*(day_end-day_start+1))
    return Ut

# Returns all processed data for one stock
def ReturnDataforSingleStock(sheet):
    
    stockname  = sheet['A1'].value
    number_of_rows = sheet.max_row 
    
    # row_offset is to offset the stock name and the Price Volume Time header rows
    row_offset = 2;    
    range_string = 'A1:' + 'L'+ str(number_of_rows-row_offset)
    
    # Time Variables to check the timestamp matches active hours
    StartTime = datetime.now().time()
    EndTime = datetime.now().time()
    EndTime_tol = datetime.now().time()
    StartTime_tol = datetime.now().time()
    
    StartTime = StartTime.replace(hour = 9, minute = 30, second =1,microsecond =0 )
    EndTime_tol = EndTime_tol.replace(hour = 15, minute = 58 , second =59, microsecond =0) 
    EndTime = EndTime.replace(hour = 15, minute = 59, second =1, microsecond =0)
    StartTime_tol = StartTime_tol.replace(hour =9, minute =29, second =59, microsecond =0)
    #Create an object of PVTforAllDays Class for one Stock
    Object_stock = PVTforAllDays(stockname)
    
    # Put the Excel data into our data structure with DateTime checks
    for row in sheet.iter_rows(range_string, row_offset=2):
        temp = row[0].value
        temp_string = temp.strftime('%H%M%S')
        current_time = datetime.strptime(temp_string,'%H%M%S').time()
        if current_time <= StartTime and current_time >= StartTime_tol:
            EachDay = PriceVolumeTime(stockname,temp.date())
        if current_time > StartTime_tol and current_time < EndTime:  
            EachDay.time.append(current_time)
            EachDay.price.append(row[1].value)
            EachDay.volume.append(row[2].value)
        if current_time > EndTime_tol and current_time < EndTime:
            Object_stock.AppendDay(EachDay)
    return Object_stock
    
    
# Main Program
print("\n"*100)
os.chdir("C:/Users/MrinalVibhav/Dropbox/ALGO/Project/DowJonesCleaned")
wb = excel.load_workbook('DJ_Cleaned.xlsx')

sheet_names = wb.get_sheet_names()

Object_stocks = []
index =0
for sheet_name in sheet_names:
    sheet = wb.get_sheet_by_name(sheet_name)
    Object_stocks.append(ReturnDataforSingleStock(sheet))  

CONST_LEN = 390
#count_rows_MAX = 390
#for index in range(len(Object_stocks)):
#    print (Object_stocks[index].NumberofDays())
#    for x in Object_stocks[index]:
#        print(x.stockname,x.date,x.Length())

# Getting the Ut Vector
Ut = []
stock_start =0
stock_end = 6
Day_start = 0
Day_end = 10
# Includes  the starting day and ending day to evaluate and same with starting stock and stock_end
Ut = ReturnUt(Object_stocks,stock_start,stock_end,Day_start,Day_end)
Mt = []
Expected_Vol = 0 
Temp_Vol =0

Actual_Vol= []
VWAP = []
for x in range(Day_end - Day_start):
    Temp_Vol += sum(Object_stocks[3].PriceVolumeTimeVectors[Day_start+x].volume)
Expected_Vol = Temp_Vol/(Day_end - Day_start + 1)

for x in range(CONST_LEN):
    Mt.append(Object_stocks[3].PriceVolumeTimeVectors[Day_end+1].volume[x])

Actual_Vol = np.cumsum(Mt)

for x in range(CONST_LEN):
    VWAP.append(Object_stocks[3].PriceVolumeTimeVectors[Day_end+1].price[x]*Mt[x])
VWAP = np.cumsum(VWAP)
for x in range(CONST_LEN):
    VWAP[x] = VWAP[x]/Actual_Vol[x]
for x in range(CONST_LEN):
    Mt[x] = Mt[x]/Expected_Vol

Y = np.array(Mt)
X = np.array(Ut)
X = np.cumsum(X)
Y = np.cumsum(Y)
#print (len(Ut),len(Mt))
#for x in range(len(Ut)):
A = pd.Series(X,index = pd.date_range("9:30","15:59",freq = "1min"))
B = pd.Series(Y,index = pd.date_range("9:30","15:59",freq = "1min"))

plt.plot(A, label = r'$\Sigma$' + ' Mt/E[V]')
plt.plot(B, label = r'$\Sigma$'+ ' Ut')
plt.xlabel('Time of the Day')
plt.ylabel('Cumulative Sum')
plt.title('Plot of ' +  r'$\Sigma$'+ 'Ut and ' + r'$\Sigma$'+ 'Mt/E[V]')
plt.legend(loc = 'upper left')
plt.show()
    
SumUt = np.cumsum(Ut)
# Plotting the VWAP vs. Order*Price
for x in range(CONST_LEN):
    Ut[x] = Ut[x]*Object_stocks[3].PriceVolumeTimeVectors[Day_end+1].price[x]
# Now Ut stores Order*Price

Y = np.array(VWAP)
X = np.array(Ut)
X = np.cumsum(X)
for x in range(CONST_LEN):
    X[x] = X[x]/SumUt[x]
#print (len(Ut),len(Mt))
#for x in range(len(Ut)):
A = pd.Series(X,index = pd.date_range("9:30","15:59",freq = "1min"))
B = pd.Series(Y,index = pd.date_range("9:30","15:59",freq = "1min"))

plt.plot(A, label = r'$\Sigma$' + ' Order*Price')
plt.plot(B, label = 'Cumulative VWAP')
plt.xlabel('Time of the Day')
plt.ylabel('Cumulative Sum')
plt.title('Plot of ' +  r'$\Sigma$'+ 'VWAP and ' + r'$\Sigma$'+ 'Order size*Price')
plt.legend(loc = 'bottom right')
plt.show()
      
        
print ('Final VWAP = ', VWAP[389])
print ('Ut*Price =', X[389])

        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    
